#!/bin/bash python3

import os
import sys
import openpyxl
from dkbc.dkbc import DKBC
from dkbc.dkbc_codes import *


dkbc = DKBC(debug=True)

def process_data(data):
    if "HttpStatusCode" in data and data["HttpStatusCode"] == 404:
        print(data["Message"])
        print(data["Details"])

    if len(data["Products"]) == 0:
        #print("Part not found...")
        data["QuantityAvailable"] = 0
        data["ProductUrl"] = 0
        return data

    #print(data["Products"])

    products = []
    for product in data["Products"]:
        if product["MinimumOrderQuantity"] > 1:
            continue

        if "Digi-Reel" in product["Packaging"]["Value"]:
            continue

        products.append(product)

    try:
        return products[0]
    except:
        data["QuantityAvailable"] = 0
        data["ProductUrl"] = 0
        return data
xlsx = openpyxl.load_workbook(sys.argv[1])
sheet = xlsx.active
rows = sheet.rows

mpn_column = 4
data_starts_on_row = 9

print('grabbing MPNs...')
for value in sheet.iter_cols(
            min_row = data_starts_on_row,
            min_col = mpn_column,
            max_col = mpn_column, 
            values_only = True):
    
    all_mpns = (list(value))
print("done.")

print('querying digikey...')
all_mpn_data = []
for mpn in all_mpns:
    x = dkbc.get_part_details(mpn)
    data = process_data(x)
    print(mpn, data["QuantityAvailable"], data["ProductUrl"])
    all_mpn_data.append(data)
print('done.')

sheet.cell(row = data_starts_on_row-1, column = sheet.max_column+1).value = 'Digikey'
print('writing quantities to sheet...')
for cur_row, data in enumerate(all_mpn_data, start = data_starts_on_row):
    sheet.cell(row = cur_row, column = sheet.max_column).value = '=HYPERLINK("{}", "{}")'.format(data["ProductUrl"], data["QuantityAvailable"])
print('done!')
xlsx.save(sys.argv[1])


